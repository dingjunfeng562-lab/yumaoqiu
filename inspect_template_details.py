import glob, json, os
import openpyxl
from openpyxl.utils import get_column_letter

p=[x for x in glob.glob(r'D:\\桌面\\比赛\\*.xlsx') if not os.path.basename(x).startswith('~$') and '应职院' in x][0]
wb=openpyxl.load_workbook(p, data_only=False)
for ws in wb.worksheets:
 print('\n###',ws.title)
 for r in range(1,ws.max_row+1):
  vals=[ws.cell(r,c).value for c in range(1,ws.max_column+1)]
  non=[f'{get_column_letter(c+1)}={v!r}' for c,v in enumerate(vals) if v is not None]
  if non: print(r,' | '.join(non))
