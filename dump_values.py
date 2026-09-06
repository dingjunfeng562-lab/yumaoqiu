import openpyxl,glob
p=glob.glob(r'D:\桌面\比赛\*.xlsx')[0]
wb=openpyxl.load_workbook(p,data_only=False)
for ws in wb.worksheets:
    print('\n###',ws.title,ws.calculate_dimension())
    for r in range(1,ws.max_row+1):
        vals=[ws.cell(r,c).value for c in range(1,ws.max_column+1)]
        if any(v is not None for v in vals):
            print(r, '|'.join('' if v is None else str(v).replace('\n','\\n') for v in vals))
