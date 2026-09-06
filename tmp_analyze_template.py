import openpyxl, glob, json, os, sys
from collections import Counter
p = glob.glob(r'D:\桌面\比赛\*.xlsx')[0]
wb = openpyxl.load_workbook(p, data_only=False)
print('PATH', repr(p))
print('SHEETS', wb.sheetnames)
print('DEFINED_NAMES', list(wb.defined_names))
for ws in wb.worksheets:
    print('\n=== SHEET', ws.title, '===')
    print('dimension', ws.calculate_dimension(), 'max', ws.max_row, ws.max_column, 'state', ws.sheet_state)
    print('freeze', ws.freeze_panes, 'filter', ws.auto_filter.ref, 'print_area', ws.print_area, 'print_titles', ws.print_title_rows, ws.print_title_cols)
    print('page_setup', {'orientation':ws.page_setup.orientation,'paperSize':ws.page_setup.paperSize,'fitToWidth':ws.page_setup.fitToWidth,'fitToHeight':ws.page_setup.fitToHeight,'scale':ws.page_setup.scale,'pageOrder':ws.page_setup.pageOrder})
    print('margins', {k:getattr(ws.page_margins,k) for k in ['left','right','top','bottom','header','footer']})
    print('sheet_props', {'gridLines':ws.sheet_view.showGridLines,'zoom':ws.sheet_view.zoomScale,'zoomNormal':ws.sheet_view.zoomScaleNormal,'defaultRowHeight':ws.sheet_format.defaultRowHeight,'defaultColWidth':ws.sheet_format.defaultColWidth})
    print('merges', [str(x) for x in ws.merged_cells.ranges])
    print('cols', [(k,round(v.width,3),v.hidden,v.bestFit,v.outlineLevel) for k,v in ws.column_dimensions.items()])
    print('rows', [(k,v.height,v.hidden,v.outlineLevel) for k,v in ws.row_dimensions.items() if v.height is not None or v.hidden or v.outlineLevel])
    print('tables', list(ws.tables), 'images',len(ws._images),'charts',len(ws._charts),'conditional',len(ws.conditional_formatting),'validations',len(ws.data_validations.dataValidation))
    print('row_breaks', [(b.id,b.min,b.max,b.man,b.pt) for b in ws.row_breaks.brk], 'col_breaks', [(b.id,b.min,b.max,b.man,b.pt) for b in ws.col_breaks.brk])
    # values and formulas in compact coordinate form
    for row in ws.iter_rows():
        vals=[]
        for c in row:
            if c.value is not None:
                vals.append(f'{c.coordinate}={c.value!r} [style={c.style_id},num={c.number_format!r}]')
        if vals: print('ROW', row[0].row, ' | '.join(vals))
    # style summary for nonempty cells and all cells in used range
    ctr=Counter(c.style_id for row in ws.iter_rows() for c in row if c.value is not None)
    print('style_counts_nonempty', dict(ctr))
    for sid in sorted(ctr):
        # find first populated cell
        first=next(c for row in ws.iter_rows() for c in row if c.value is not None and c.style_id==sid)
        print('STYLE',sid,'first',first.coordinate,'font',first.font.name,first.font.sz,first.font.bold,first.font.italic,first.font.color.type if first.font.color else None, first.font.color.rgb if first.font.color and first.font.color.type=='rgb' else None,'fill',first.fill.fill_type,first.fill.fgColor.type,first.fill.fgColor.rgb,'align',first.alignment.horizontal,first.alignment.vertical,first.alignment.wrap_text,'border',(first.border.left.style,first.border.right.style,first.border.top.style,first.border.bottom.style))
