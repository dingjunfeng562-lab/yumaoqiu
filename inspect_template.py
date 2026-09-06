import glob, json, os
import openpyxl

paths = glob.glob(r'D:\\桌面\\比赛\\*.xlsx')
print('FILES', json.dumps(paths, ensure_ascii=False))
p = [x for x in paths if os.path.basename(x).startswith('应职院')][0]
wb = openpyxl.load_workbook(p, read_only=False, data_only=False)
print('SHEETS', json.dumps(wb.sheetnames, ensure_ascii=False))
for ws in wb.worksheets:
    print('SHEET', json.dumps({'title': ws.title, 'rows': ws.max_row, 'cols': ws.max_column}, ensure_ascii=False))
    print('MERGES', list(ws.merged_cells.ranges))
    print('WIDTHS', {k: v.width for k,v in ws.column_dimensions.items() if v.width})
    print('HEIGHTS', {k: v.height for k,v in ws.row_dimensions.items() if v.height})
    print('FREEZE', ws.freeze_panes)
    print('PRINT', ws.print_area, ws.page_setup.orientation, ws.page_setup.paperSize, ws.sheet_properties.pageSetUpPr.fitToPage if ws.sheet_properties.pageSetUpPr else None)
    print('HIDDEN', ws.sheet_state)
    print('FORMULAS', [(c.coordinate,c.value) for row in ws.iter_rows() for c in row if isinstance(c.value,str) and c.value.startswith('=')][:20])
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 50), values_only=True):
        print(json.dumps(list(row[:min(ws.max_column, 20)]), ensure_ascii=False))
