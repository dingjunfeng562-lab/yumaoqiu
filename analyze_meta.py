import openpyxl, glob, json
p=glob.glob(r'D:\桌面\比赛\*.xlsx')[0]
wbF=openpyxl.load_workbook(p,data_only=False)
wbV=openpyxl.load_workbook(p,data_only=True)
out={'path':p,'sheets':[],'defined_names':list(wbF.defined_names)}
for ws,wv in zip(wbF.worksheets,wbV.worksheets):
    d={'title':ws.title,'dimension':ws.calculate_dimension(),'max_row':ws.max_row,'max_col':ws.max_column,'state':ws.sheet_state,'freeze':str(ws.freeze_panes),'merges':[str(x) for x in ws.merged_cells.ranges], 'print_area':str(ws.print_area), 'print_titles_rows':str(ws.print_title_rows),'print_titles_cols':str(ws.print_title_cols), 'page_setup':{k:getattr(ws.page_setup,k) for k in ['orientation','paperSize','fitToWidth','fitToHeight','scale','pageOrder']}, 'margins':{k:getattr(ws.page_margins,k) for k in ['left','right','top','bottom','header','footer']}, 'view':{'grid':ws.sheet_view.showGridLines,'zoom':ws.sheet_view.zoomScale,'zoomNormal':ws.sheet_view.zoomScaleNormal}, 'default_row':ws.sheet_format.defaultRowHeight,'default_col':ws.sheet_format.defaultColWidth, 'cols':{}, 'rows':{}, 'formula_cells':[], 'nonempty':[], 'cached_formulas':[]}
    for k,v in ws.column_dimensions.items():
        d['cols'][k]={'width':v.width,'hidden':v.hidden,'bestFit':v.bestFit,'outline':v.outlineLevel}
    for k,v in ws.row_dimensions.items():
        if v.height is not None or v.hidden or v.outlineLevel:
            d['rows'][str(k)]={'height':v.height,'hidden':v.hidden,'outline':v.outlineLevel}
    for row in ws.iter_rows():
        for c in row:
            if c.value is not None:
                ent={'coord':c.coordinate,'value':c.value,'style':c.style_id,'numfmt':c.number_format}
                d['nonempty'].append(ent)
                if isinstance(c.value,str) and c.value.startswith('='):
                    d['formula_cells'].append(ent)
                    d['cached_formulas'].append({'coord':c.coordinate,'formula':c.value,'cached':wv[c.coordinate].value})
    d['tables']=list(ws.tables); d['images']=len(ws._images); d['charts']=len(ws._charts); d['conditional']=len(ws.conditional_formatting); d['validations']=len(ws.data_validations.dataValidation)
    d['row_breaks']=[{'id':b.id,'min':b.min,'max':b.max,'man':b.man,'pt':b.pt} for b in ws.row_breaks.brk]
    d['col_breaks']=[{'id':b.id,'min':b.min,'max':b.max,'man':b.man,'pt':b.pt} for b in ws.col_breaks.brk]
    out['sheets'].append(d)
json.dump(out,open('template_meta.json','w',encoding='utf-8'),ensure_ascii=False,indent=2,default=str)
